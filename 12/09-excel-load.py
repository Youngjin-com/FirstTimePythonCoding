import openpyxl

filename = "./12/excel_sample.xlsx"
workbook = openpyxl.load_workbook(filename)
sheet = workbook.active
for row in range(1, sheet.max_row):
    for col in range(1, sheet.max_column):
        data = sheet.cell(row, col).value
        print(data, end=" ")
    print()